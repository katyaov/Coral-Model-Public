from typing import Tuple, Any, Dict
import builtins
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd


# -------------------------------------------------------
# Safe exec to evaluate user_inputs_old_X.py
# -------------------------------------------------------
def _safe_exec_py(path: str) -> Dict[str, Any]:
    path = Path(path)
    src = path.read_text()

    # Allow only a safe subset of builtins
    allowed_builtins = {
        "True": True,
        "False": False,
        "None": None,
        "dict": dict,
        "list": list,
        "tuple": tuple,
        "set": set,
        "range": range,
        "len": len,
        "sum": sum,   # <--- important for your error
        "min": min,
        "max": max,
        "abs": abs,
    }

    safe_globals: Dict[str, Any] = {"__builtins__": allowed_builtins}
    safe_locals: Dict[str, Any] = {}

    exec(compile(src, str(path), "exec"), safe_globals, safe_locals)

    # Merge globals + locals (locals win)
    ns = {**safe_globals, **safe_locals}
    return ns


# -------------------------------------------------------
# Conversion helper
# -------------------------------------------------------
def _get(ns: Dict[str, Any], keys, default=None):
    for k in keys:
        if k in ns:
            return ns[k]
    return default


# -------------------------------------------------------
# Main convert function
# -------------------------------------------------------
def convert(
    old_py_path: str,
    template_excel_path: str,
    out_excel_path: str,
    out_user_input_path: str
) -> Tuple[str, str]:
    """
    Converts old user_inputs file into:
      - A new Excel file populated with parameters (and real data if provided)
      - A new user_inputs python file
    Returns: (out_excel_path, out_user_input_path)
    """

    # 1. Load the namespace from the old python file
    ns = _safe_exec_py(old_py_path)

    # 2. Load template workbook
    wb = load_workbook(template_excel_path)

    # 3. Populate Real_Cover sheet if real_data dict is present
    real_data = ns.get("real_data")
    if real_data is not None:
        df_real = pd.DataFrame(real_data)

        if "Real_Cover" not in wb.sheetnames:
            raise ValueError("Template must have a sheet called 'Real_Cover'")

        ws = wb["Real_Cover"]

        # Clear everything in Real_Cover
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # Write headers
        for j, col in enumerate(df_real.columns, start=1):
            ws.cell(row=1, column=j, value=col)

        # Write values
        for i, row in df_real.iterrows():
            for j, val in enumerate(row, start=1):
                ws.cell(row=i+2, column=j, value=val)

    # 4. Save workbook
    wb.save(out_excel_path)

    # 5. Write out new user_inputs file (basic dump of variables)
    with open(out_user_input_path, "w") as f:
        f.write("# Auto-generated from convert_coral_inputs.py\n\n")
        for k, v in ns.items():
            if not k.startswith("__"):  # skip internals
                f.write(f"{k} = {repr(v)}\n")

    return out_excel_path, out_user_input_path
